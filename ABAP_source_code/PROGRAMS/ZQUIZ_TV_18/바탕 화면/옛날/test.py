import openpyxl as op
import os

CH_name = list()
PAST_name = list()
upload_date =list()

path = 'C:\강민구\\2021년\책가방보내기_2021입학년도\\4. 선정 준비\선정 작업중\증명서류\\file\\2021년'


wb = op.load_workbook(path+'\책가방예시.xlsx')
ws = wb["책가방"]


file_name_li = os.listdir(path)

cullA = ws["A"]
for cell in cullA :
    CH_name.append(cell.value)

cullB = ws["B"]
for cell in cullB :
    PAST_name.append(cell.value)

cullC = ws["C"]
for cell in cullC :
    upload_date.append(cell.value)

